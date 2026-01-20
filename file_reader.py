import os

class FileReader:
    def read_text_file(self, file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    return file.read()
            except:
                return ""
        except:
            return ""
    
    def read_pdf_file(self, file_path):
        text = ""
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except:
            try:
                text = self.read_text_file(file_path)
            except:
                text = ""
        return text
    
    def read_docx_file(self, file_path):
        text = ""
        try:
            from docx import Document
            doc = Document(file_path)
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
        except:
            try:
                text = self.read_text_file(file_path)
            except:
                text = ""
        return text
    
    def read_excel_file(self, file_path):
        text = ""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = ' '.join([str(cell) for cell in row if cell is not None])
                    text += row_text + "\n"
        except:
            try:
                text = self.read_text_file(file_path)
            except:
                text = ""
        return text
    
    def read_file(self, file_path):
        filename = os.path.basename(file_path)
        extension = os.path.splitext(filename)[1].lower()
        
        if extension == '.txt':
            return self.read_text_file(file_path)
        elif extension == '.pdf':
            return self.read_pdf_file(file_path)
        elif extension in ['.docx', '.doc']:
            return self.read_docx_file(file_path)
        elif extension in ['.xlsx', '.xls']:
            return self.read_excel_file(file_path)
        else:
            return self.read_text_file(file_path)